import tkinter as tk
from datetime import datetime, timedelta
from decimal import Decimal
from tkinter import messagebox, ttk, filedialog
from tktimepicker import AnalogPicker, AnalogThemes

import openpyxl
from openpyxl.styles import Font, PatternFill
from tkcalendar import Calendar, DateEntry
from zk import ZK


def range_date(attendaces, start_date, end_date):
    result = []
    #  datetime.date to datetime.datetime
    start_date = datetime.combine(start_date, datetime.min.time())
    end_date = datetime.combine(end_date, datetime.max.time())
    for attendance in attendaces:
        if attendance.timestamp >= start_date and attendance.timestamp <= end_date:
            result.append(attendance)
    return result


def blow_it(
    start_date, end_date, ip, data_file_location, shift_length, shift_start_time
):
    conn = None
    zk = ZK(
        ip,
        port=4370,
        timeout=1,
        password=0,
        force_udp=False,
        ommit_ping=False,
    )
    dict = {}
    attendances = []
    try:
        conn = zk.connect()
        conn.disable_device()
        attendances = conn.get_attendance()
        conn.enable_device()
    except Exception as e:
        print("Process terminate : {}".format(e))
        return False
    finally:
        if conn:
            conn.disconnect()

    final = range_date(attendances, start_date, end_date)
    for i in final:
        if i.user_id in dict:
            dict[i.user_id].append(i.timestamp)
        else:
            dict[i.user_id] = []
            dict[i.user_id].append(i.timestamp)

    wb = openpyxl.load_workbook(data_file_location, data_only=True)
    sheet = wb.active
    sheet.sheet_view.rightToLeft = True
    Emplyees = {}

    for row in sheet.iter_rows(min_row=2):
        if row[1].value == None:
            continue
        EmpId = row[0].value
        EmpName = row[1].value
        EmpSalary = row[2].value
        EmpSalaryPerHour = row[3].value
        if str(EmpId) in dict:
            Emplyees[EmpId] = [EmpName, Decimal(EmpSalary), Decimal(EmpSalaryPerHour)]

    for id in dict:
        Intid = int(id)
        if Intid in Emplyees:
            Emplyees[Intid].append(dict[id])

    wb = openpyxl.Workbook()
    sheet = wb.active
    # change sheet direction
    sheet.sheet_view.rightToLeft = True
    # change sheet title
    sheet.title = "المرتبات"
    sheet.sheet_properties.tabColor = "1072BA"

    sheet.append(["الكود", "الاسم", "عدد ساعات العمل", "سعر الساعة", "اجمالي الراتب"])
    FontObj = Font(name="Arial", size=16, bold=True, italic=False, color="FF0000")
    sheet["A1"].font = FontObj
    sheet["B1"].font = FontObj
    sheet["C1"].font = FontObj
    sheet["D1"].font = FontObj
    sheet["E1"].font = FontObj
    # modify each cell size
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 30
    sheet.column_dimensions["E"].width = 30
    # change background color for first row to gray
    sheet["A1"].fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    sheet["B1"].fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    sheet["C1"].fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    sheet["D1"].fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )
    sheet["E1"].fill = PatternFill(
        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
    )

    FontObjforData = Font(
        name="Arial", size=14, bold=False, italic=False, color="000000"
    )
    FontForSum = Font(name="Arial", size=14, bold=True, italic=False, color="000000")
    for id in Emplyees:
        secondPrice = (Emplyees[id][2] / 60) / 60
        timeSpent = 0
        for i in range(0, len(Emplyees[id][3]) - 1, 2):
            daySec = (Emplyees[id][3][i + 1] - Emplyees[id][3][i]).total_seconds()
            timeSpent += daySec
        sheet.append(
            [
                id,
                Emplyees[id][0],
                (timeSpent / 60) / 60,
                Emplyees[id][2],
                Decimal(timeSpent) * Decimal(secondPrice),
            ]
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = FontObjforData

    for id in Emplyees:
        wb.create_sheet(title=str(id))
        sheet = wb[str(id)]
        sheet.sheet_view.rightToLeft = True
        # سعر اليوم - سعر الاضافي - الوقت الاضافي - الاجمالي
        sheet.append(
            [
                "الكود",
                "الاسم",
                "حضور",
                "انصراف",
                "الوقت",
                "سعر الساعة",
                "اجمالي اليوم",
                "الاضافي",
                "اجمالي الاضافي",
                "الملاحظات",
            ]
        )

        sheet.append(
            # id,name,attendance,leave,time,price,total,extra,extraTotal,notes
            [
                id,  # id
                Emplyees[id][0],  # name
                "",  # attendance
                "",  # leave
                "",  # time
                Emplyees[id][2],  # price
                Decimal(int(Emplyees[id][2])) * Decimal(shift_length),  # total
                "",  # extra
                "",  # extraTotal
                "",  # notes
            ]
        )
        timeSpent = 0
        totalRow = 0
        for i in range(0, len(Emplyees[id][3]) - 1, 2):
            # we want to calculate the if he exceed the start time by 20 min we will subtract 2 hours
            # if he exceed time by 30 min we will subtract 4 hours
            # start = 8:10 am if he exceed 8:20 am we will subtract 2 hours
            # if he exceed 8:20 am we will subtract 4 hours
            # if he came before 8:10 am and after 8:00 am we will not subtract any thing
            notes = ""

            shift_start_datetime = datetime.combine(datetime.today(), shift_start_time)
            shift_end_datetime = shift_start_datetime + timedelta(
                hours=int(shift_length)
            )
            daySec = (
                Emplyees[id][3][i + 1] - min(Emplyees[id][3][i], shift_end_datetime)
            ).total_seconds()

            if (
                Emplyees[id][3][i].time()
                > (shift_start_datetime + timedelta(minutes=10)).time()
                and Emplyees[id][3][i].time()
                < (shift_start_datetime + timedelta(minutes=20)).time()
            ):
                daySec -= 2 * 60 * 60
                notes = "تأخير اكثر من 10 دقائق"
            elif (
                Emplyees[id][3][i].time()
                > (shift_start_datetime + timedelta(minutes=20)).time()
            ):
                daySec -= 4 * 60 * 60
                notes = "تأخير اكثر من 20 دقيقة"

            # Assuming shift_length is an integer representing hours
            shift_end_time = (
                shift_start_datetime + timedelta(hours=int(shift_length))
            ).time()

            # Assuming Emplyees[id][3][i + 1] is a datetime.datetime object
            overTime = (
                datetime.combine(datetime.today(), Emplyees[id][3][i + 1].time())
                - datetime.combine(datetime.today(), shift_end_time)
            ).total_seconds()

            overTime = max(overTime, 0)

            if overTime:
                notes += "\n عمل وقت اضافي" + " " + str((overTime / 60) / 60)

            daySec = max(daySec, 0)
            timeSpent += daySec
            secondPrice = (Emplyees[id][2] / 60) / 60
            sheet.append(
                [
                    "",  # id
                    "",  # name
                    Emplyees[id][3][i],  # attendance
                    Emplyees[id][3][i + 1],  # leave
                    (daySec / 60) / 60,  # time
                    Emplyees[id][2],  # price
                    Decimal(daySec) * Decimal(secondPrice),  # total
                    (overTime / 60) / 60,  # extra
                    Decimal(overTime) * Decimal(secondPrice),  # extraTotal
                    notes,  # notes
                ]
            )
            totalRow += 1
        # sheet.append(
        #     ["", "", "", "", "", "", Decimal(timeSpent) * Decimal(secondPrice), ""]
        # )
        # sheet["G" + str(totalRow + 3)].font = FontForSum
        # sheet["G" + str(totalRow + 3)].fill = PatternFill(
        #     start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        # )
        colered_cell = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1"]
        for cell in colered_cell:
            sheet[cell].font = FontObj
            sheet[cell].fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 30
        sheet.column_dimensions["D"].width = 30
        sheet.column_dimensions["E"].width = 30
        sheet.column_dimensions["F"].width = 30
        sheet.column_dimensions["G"].width = 30
        sheet.column_dimensions["H"].width = 30
        sheet.column_dimensions["I"].width = 30
        sheet.column_dimensions["J"].width = 50

    start_time = start_date.strftime("%Y-%m-%d")
    end_time = end_date.strftime("%Y-%m-%d")
    wb.save("Data/" + start_time + "+" + end_time + ".xlsx")
    return True


def submit_dates():
    ip = ip_entry.get()
    start_date = start_calendar.get_date()
    end_date = end_calendar.get_date()
    data_file_location = file_location.get()
    shift_length = shift_entry.get()
    time = time_picker.time()
    hrs = time[0]
    mins = time[1]
    event = time[2]
    time_to_datetime = datetime.strptime(f"{hrs}:{mins} {event}", "%I:%M %p")
    print(time_to_datetime.time())

    isDone = blow_it(
        start_date,
        end_date,
        ip,
        data_file_location,
        # to int because it's string
        int(shift_length),
        time_to_datetime.time(),
    )
    if not isDone:
        messagebox.showerror("خطأ", "الرجاء التأكد من الاتصال بالجهاز !")
        return
    if start_date and end_date and isDone:
        messagebox.showinfo(
            "تم عمل التقرير", f"تاريخ البدء: {start_date}\nتاريخ الانتهاء: {end_date}"
        )
    else:
        messagebox.showerror("خطأ", "الرجاء تحديد تاريخ البدء وتاريخ الانتهاء!")


root = tk.Tk()
root.title("تقرير الحضور والانصراف للشركة المصرية")

ip_label = ttk.Label(root, text=": ادخل عنوان الجهاز")
ip_label.grid(row=0, column=1, padx=0, pady=0, sticky="w")
ip_entry = ttk.Entry(root)
ip_entry.grid(row=0, column=0, padx=0, pady=0)
ip_entry.insert(0, "192.168.1.202")

# Calendar start date and end date
start_label = ttk.Label(root, text=": تاريخ البدء")
end_label = ttk.Label(root, text=": تاريخ الانتهاء")
start_label.grid(row=2, column=1, padx=10, pady=10, sticky="w")
end_label.grid(row=3, column=1, padx=10, pady=10, sticky="w")

start_calendar = DateEntry(root, date_pattern="yyyy-MM-dd", selectmode="day")
end_calendar = DateEntry(root, date_pattern="yyyy-MM-dd", selectmode="day")

start_calendar.grid(row=2, column=0, padx=10, pady=10)
end_calendar.grid(row=3, column=0, padx=10, pady=10)

# choose file button
file_location = ttk.Entry(root, width=50)
file_location.grid(row=4, column=0, padx=50, pady=5, sticky="w")


def open_text_file():
    filetypes = (("Excel files", "*.xlsx"), ("All files", "*.*"))
    f = filedialog.askopenfile(filetypes=filetypes, initialdir="D:/Downloads")
    file_location.insert("0", f.name)


choose_file_button = ttk.Button(root, text="اختر ملف", command=open_text_file)
choose_file_button.grid(row=4, column=1, columnspan=2, padx=10, pady=20)

# shift start time
shift_start_label = ttk.Label(root, text=": بداية الوردية")
shift_start_label.grid(row=5, column=1, padx=10, pady=10, sticky="w")
time_picker = AnalogPicker(root)
time_picker.grid(row=5, column=0, padx=10, pady=10, sticky="w")
theme = AnalogThemes(time_picker)
theme.setNavyBlue()
time = time_picker.time()


# shift length
shift_label = ttk.Label(root, text=": مدة الوردية")
shift_label.grid(row=6, column=1, padx=10, pady=10, sticky="w")
shift_entry = ttk.Entry(root)
shift_entry.grid(row=6, column=0, padx=10, pady=10)
shift_entry.insert(0, "12")

submit_button = ttk.Button(root, text="عمل التقرير", command=submit_dates)
submit_button.grid(row=7, column=0, columnspan=2, padx=10, pady=20)

root.mainloop()
